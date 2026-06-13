import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.drawing.image import Image # RE-AJOUTÉ ICI

def _creation_du_dossier(output_path):
    # on créé le dossier
    os.makedirs(os.path.dirname(output_path), exist_ok=True)


def _onglet_data(df_raw, output_path):
    # mon onglet data dans le fichier excel
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_raw.to_excel(writer, sheet_name="DATA", index=False)


def _ouverture_du_fichier(output_path):
    # ouvertur du fichier
    return load_workbook(output_path)


def _indicateurs(wb):
    # l'onglet indicator ou je vais mettre mes graph
    ws = wb.create_sheet("Indicateurs")

    # le titre du rapport en M2 pour qu'il soit centrer 
    ws["M2"] = "TABLEAU DE BORD DU RÉSEAU VÉLIB"
    ws["M2"].font = Font(bold=True, size=16, color="1F497D")
    
    return ws


def _liste_deroulante_communes(ws, df_raw):
    # on met le filtre en B1
    ws["A1"] = "Choisir une commune :"
    ws["A1"].font = Font(bold=True, color="0000FF") # titre du filtre 
    ws["B1"] = "Paris" # valeur par défaut obligatoire

    # extraction des communes uniques pour éviter les doublons 
    communes_uniques = sorted(df_raw['nom_arrondissement_communes'].dropna().unique())
    
    # chaine de caractére pour l'affichage
    liste_communes_txt = f'"{",".join(communes_uniques)}"'
    
    # la liste deroulante
    dv = DataValidation(type="list", formula1=liste_communes_txt, allow_blank=True)
    
    # on ajoute la validation à la feuille PUIS on l'associe à la cellule
    ws.add_data_validation(dv)
    dv.add(ws["B1"])


def _recup_colonne(df_raw):
    # je recupere les colonnes 
    headers = list(df_raw.columns)

    col_capacity = headers.index("capacity") + 1
    col_docks = headers.index("numdocksavailable") + 1
    col_bikes = headers.index("numbikesavailable") + 1
    col_meca = headers.index("mechanical") + 1
    col_ebike = headers.index("ebike") + 1

    return {
        "capacity": get_column_letter(col_capacity),
        "docks": get_column_letter(col_docks),
        "bikes": get_column_letter(col_bikes),
        "meca": get_column_letter(col_meca),
        "ebike": get_column_letter(col_ebike)
    } ### dictionnaire contenant le nom de la variable et la colonne dans laquelle elle se trouve


def _kpi_du_reseau(ws, cols, max_row):
    #  indicateurs réseau velib
    ws["A4"] = "Indicateurs Réseau"
    ws["A4"].font = Font(bold=True, underline="single")

    # on aplique le calcul en fonction de la variable dans B1 (le filtre)
    ws["A5"] = "Capacité totale"
    ws["B5"] = f'=SUMIF(DATA!M2:M{max_row}, B1, DATA!{cols["capacity"]}2:{cols["capacity"]}{max_row})'

    ws["A6"] = "Bornettes libres"
    ws["B6"] = f'=SUMIF(DATA!M2:M{max_row}, B1, DATA!{cols["docks"]}2:{cols["docks"]}{max_row})'

    ws["A7"] = "Total vélos disponibles"
    ws["B7"] = f'=SUMIF(DATA!M2:M{max_row}, B1, DATA!{cols["bikes"]}2:{cols["bikes"]}{max_row})'

    ws["A8"] = "Taux d'occupation global"
    ws["B8"] = "=B7/B5"
    ws["B8"].number_format = "0.00%"

    
def _kpi_du_types_velos(ws, cols, max_row):
    # mes indicateurs sur les types de vélos
    ws["A11"] = "Répartition par type de vélo"
    ws["A11"].font = Font(bold=True, underline="single")

    ws["A12"] = "Vélos mécaniques"
    ws["B12"] = f'=SUMIF(DATA!M2:M{max_row}, B1, DATA!{cols["meca"]}2:{cols["meca"]}{max_row})'

    ws["A13"] = "Vélos électriques"
    ws["B13"] = f'=SUMIF(DATA!M2:M{max_row}, B1, DATA!{cols["ebike"]}2:{cols["ebike"]}{max_row})'

    ws["A14"] = "Total vélos vérifié"
    ws["B14"] = "=SUM(B12:B13)"


def _top5_communes(ws, df_top5):
    # top 5 des communes 
    ws["A17"] = "Top 5 des communes les plus desservies en vélo"
    ws["A17"].font = Font(bold=True, underline="single")

    # On utilise les données calculées par commune_analysis
    for idx, row in df_top5.iterrows():
        ligne_actuelle = 18 + idx
        ws["A" + str(ligne_actuelle)] = row['nom_arrondissement_communes']
        ws["B" + str(ligne_actuelle)] = round(row['numbikesavailable'], 1)
        ws["C" + str(ligne_actuelle)] = round(row['occupation_rate'] * 100, 1)


def _graphiques(ws):
    # Graphique en secteur
    pie = PieChart()
    pie.title = "Types de vélos disponibles"

    data_pie = Reference(ws, min_col=2, min_row=12, max_row=13)
    labels_pie = Reference(ws, min_col=1, min_row=12, max_row=13)

    pie.add_data(data_pie)
    pie.set_categories(labels_pie)

    ws.add_chart(pie, "E4")

    # barplot de disponibilité (remis d'origine sur les lignes 5 à 7)
    bar = BarChart()
    bar.title = "Disponibilité globale du réseau"
    bar.y_axis.title = "Quantité"
    bar.width = 15  
    bar.height = 10

    data_bar = Reference(ws, min_col=2, min_row=5, max_row=7)
    labels_bar = Reference(ws, min_col=1, min_row=5, max_row=7)

    bar.add_data(data_bar)
    bar.set_categories(labels_bar)
    bar.legend = None

    ws.add_chart(bar, "N4")

    # Graphique en barre combiné 
    chart_velo = BarChart()
    chart_velo.title = "Nombre de vélo par occupation dans chaque commune du top"
    chart_velo.style = 10 
    chart_velo.width = 15   
    chart_velo.height = 12  
    chart_velo.y_axis.title = "Nombre moyen de vélos"
    chart_velo.x_axis.title = "Communes"

    data_velo = Reference(ws, min_col=2, min_row=18, max_row=22)
    labels_communes = Reference(ws, min_col=1, min_row=18, max_row=22)
    chart_velo.add_data(data_velo)
    chart_velo.set_categories(labels_communes)
    chart_velo.legend = None

    # création de la ligne du taux d'occupat°
    chart_line = LineChart()
    data_occupation = Reference(ws, min_col=3, min_row=18, max_row=22)
    chart_line.add_data(data_occupation)
    
    # Concatenation des 2 graphiques 
    chart_line.y_axis.axId = 200
    chart_line.y_axis.title = "Taux d'occupation en (%)"
    chart_line.y_axis.crosses = "max"
    chart_velo += chart_line

    ws.add_chart(chart_velo, "E18")


def _inserer_carte(ws, image_path="data/carte_velib.png"):
    # On intègre l'image 
    if os.path.exists(image_path):
        img = Image(image_path)
        img.width = 550   
        img.height = 400  
        
        # on colle l'image en N26
        ws.add_image(img, "N26") 
        print("La capture de la carte a bien été déplacée plus bas !")
    else:
        print(f"Note : ou est ma carte {image_path}, le rapport est généré sans image.")


def _sauvegarde(wb, output_path):
    wb.save(output_path)
    wb.close()
    print("le fichier excel est op :", output_path)


# La fonction principale 
def generate_excel_report(df_raw, df_top5, output_path="data/mon_rapport_final.xlsx"):
    _creation_du_dossier(output_path)
    _onglet_data(df_raw, output_path)
    
    wb = _ouverture_du_fichier(output_path)
    max_row = len(df_raw) + 1

    ws = _indicateurs(wb)
    _liste_deroulante_communes(ws, df_raw)
    cols = _recup_colonne(df_raw)
    
    _kpi_du_reseau(ws, cols, max_row)
    _kpi_du_types_velos(ws, cols, max_row)
    _top5_communes(ws, df_top5)
    _graphiques(ws)
    
    _inserer_carte(ws, "data/carte_velib.png") # insertion de la carte
    
    _sauvegarde(wb, output_path)